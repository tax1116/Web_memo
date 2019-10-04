import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 타겟 URL을 읽어서 HTML를 받아오고,
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909', headers=headers)

soup = BeautifulSoup
# HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
# soup이라는 변수에 "파싱 용이해진 html"이 담긴 상태가 됨
# 이제 코딩을 통해 필요한 부분을 추출하면 된다.

# HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
soup = BeautifulSoup(data.text, 'html.parser')

# select를 이용해서, tr들을 불러오기
movies = soup.select('#old_content > table > tbody > tr')

movie_table = load_workbook('prac01.xlsx')
movie_sheet = movie_table['prac']

# movies (tr들) 의 반복문을 돌리기
row = 2
num = 1
for movie in movies:
    # movie 안에 a 가 있으면,
    a_tag = movie.select_one('td.title > div > a')
    if a_tag is not None:
        # a의 text를 찍어본다.
        movie_title = a_tag.text
        point = movie.select('td.point')[0].text

        movie_sheet.cell(row=row, column=1, value=num)
        movie_sheet.cell(row=row, column=2, value=movie_title)
        movie_sheet.cell(row=row, column=3, value=point)
        row += 1
        num += 1

    movie_table.save('movie.xlsx')
